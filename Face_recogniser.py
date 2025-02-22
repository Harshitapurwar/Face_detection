
import tkinter as tk
from tkinter import messagebox
import cv2
import os
import time
import numpy as np
from PIL import Image
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Excel File Setup
attendance_file = "attendance.xlsx"

def setup_excel():
    if not os.path.exists(attendance_file):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        sheet.append(["Roll Number", "Name", "Branch", "Attendance", "Timestamp"])
        workbook.save(attendance_file)

# GUI Setup
window = tk.Tk()
window.title("Face Recognition Attendance System")

# Labels and Entry boxes for Name, Roll Number, and Branch
tk.Label(window, text="Name:", font=("Arial", 14)).grid(row=0, column=0, padx=10, pady=5)
entry_name = tk.Entry(window, font=("Arial", 14), width=30)
entry_name.grid(row=0, column=1, padx=10, pady=5)

tk.Label(window, text="Roll Number:", font=("Arial", 14)).grid(row=1, column=0, padx=10, pady=5)
entry_roll = tk.Entry(window, font=("Arial", 14), width=30)
entry_roll.grid(row=1, column=1, padx=10, pady=5)

tk.Label(window, text="Branch:", font=("Arial", 14)).grid(row=2, column=0, padx=10, pady=5)
entry_branch = tk.Entry(window, font=("Arial", 14), width=30)
entry_branch.grid(row=2, column=1, padx=10, pady=5)

# Dataset generation function


def generate_dataset():
    user_id = entry_roll.get()
    name = entry_name.get()
    branch = entry_branch.get()

    if not user_id or not name or not branch:
        messagebox.showinfo("Error", "Please fill all details.")
        return

    # Load the trained classifier if available
    if os.path.exists("classifier.xml"):
        clf = cv2.face.LBPHFaceRecognizer_create()
        clf.read("classifier.xml")

    face_classifier = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")

    def face_cropped(img):
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = face_classifier.detectMultiScale(gray, 1.3, 5)
        if len(faces) == 0:
            return None
        for (x, y, w, h) in faces:
            return img[y:y + h, x:x + w]

    os.makedirs("data", exist_ok=True)
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        messagebox.showerror("Error", "Could not open video source.")
        return

    img_id = 0
    while True:
        ret, frame = cap.read()
        if not ret:
            break
        cropped_face = face_cropped(frame)
        if cropped_face is not None:
            img_id += 1
            face = cv2.resize(cropped_face, (200, 200))
            face = cv2.cvtColor(face, cv2.COLOR_BGR2GRAY)

            # Check if face matches any existing dataset
            if os.path.exists("classifier.xml"):
                id, confidence = clf.predict(face)
                if confidence < 50:  # Low confidence means match found
                    messagebox.showwarning(
                        "Warning",
                        f"This face is already associated with Roll Number: {id}. "
                        "Please use the same ID or contact the admin.",
                    )
                    cap.release()
                    cv2.destroyAllWindows()
                    return

            file_name_path = f"data/user.{user_id}.{img_id}.jpg"
            cv2.imwrite(file_name_path, face)
            cv2.imshow("Cropped Face", face)

        if cv2.waitKey(1) == 13 or img_id == 100:  # Stop when 100 images are captured
            break

    cap.release()
    cv2.destroyAllWindows()
    messagebox.showinfo("Result", f"Dataset for {name} created successfully.")

# Training the classifier function
def train_classifier():
    data_dir = "data"
    path = [os.path.join(data_dir, f) for f in os.listdir(data_dir)]
    faces, ids = [], []
    for image in path:
        img = Image.open(image).convert('L')
        imageNp = np.array(img, 'uint8')
        id = int(os.path.split(image)[1].split(".")[1])
        faces.append(imageNp)
        ids.append(id)

    ids = np.array(ids)
    clf = cv2.face.LBPHFaceRecognizer_create()
    clf.train(faces, ids)
    clf.write("classifier.xml")
    messagebox.showinfo("Result", "Training completed successfully.")

# Mark Attendance in Excel
def mark_attendance(roll_number, name, branch, status):
    workbook = load_workbook(attendance_file)
    sheet = workbook.active
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[0].value == roll_number:
            row[3].value = status
            row[4].value = timestamp
            workbook.save(attendance_file)
            return
    
    sheet.append([roll_number, name, branch, status, timestamp])
    workbook.save(attendance_file)

# Recognizing faces
def recognize_face():
    faceCascade = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")
    clf = cv2.face.LBPHFaceRecognizer_create()
    clf.read("classifier.xml")

    recognized_rolls = set()

    cap = cv2.VideoCapture(0)
    while True:
        ret, frame = cap.read()
        if not ret:
            break
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = faceCascade.detectMultiScale(gray, 1.3, 5)
        for (x, y, w, h) in faces:
            id, confidence = clf.predict(gray[y:y + h, x:x + w])
            if confidence < 70:
                roll_number = str(id)
                recognized_rolls.add(roll_number)
                mark_attendance(roll_number, entry_name.get(), entry_branch.get(), "Present")
                cv2.putText(frame, f"ID: {roll_number}", (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 2)
            else:
                cv2.putText(frame, "Unknown", (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)
        cv2.imshow("Face Recognition", frame)
        if cv2.waitKey(1) == 13:  # Press Enter to exit
            break

    cap.release()
    cv2.destroyAllWindows()

    # Mark absent students after 10 minutes
    time.sleep(600)
    workbook = load_workbook(attendance_file)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[0].value not in recognized_rolls:
            row[3].value = "Absent"
            row[4].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    workbook.save(attendance_file)

# Buttons for actions
tk.Button(window, text="Generate Dataset", font=("Arial", 14), command=generate_dataset, bg="lightblue").grid(row=3, column=0, columnspan=2, pady=10)
tk.Button(window, text="Train Classifier", font=("Arial", 14), command=train_classifier, bg="lightgreen").grid(row=4, column=0, columnspan=2, pady=10)
tk.Button(window, text="Recognize Face", font=("Arial", 14), command=recognize_face, bg="lightcoral").grid(row=5, column=0, columnspan=2, pady=10)

setup_excel()
window.geometry("500x400")
window.mainloop()
